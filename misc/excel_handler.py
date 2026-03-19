# hello.py

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main():
    # 1) 간단한 인사 출력
    print("Hello, World!")

    # 2) 엑셀 파일 열기
    wb = load_workbook("test.xlsx")
    ws = wb.active  # 첫 번째 시트 사용

    # 3) 셀 배경색 정의 (ARGB 포맷)
    green_fill  = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")  # 녹색
    orange_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")  # 주황색

    # 4) 셀 순회하며 값에 따라 채우기
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value == "PASS":
                cell.fill = green_fill
            elif cell.value == "FAIL":
                cell.fill = orange_fill

    # 5) 변경사항 저장 (원본 덮어쓰려면 동일한 이름으로, 아니면 다른 이름으로)
    wb.save("test_colored.xlsx")
    print("✅ Coloring 완료! 결과를 test_colored.xlsx로 저장했습니다.")

if __name__ == "__main__":
    main()
